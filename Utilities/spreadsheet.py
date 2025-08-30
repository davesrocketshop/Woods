# ***************************************************************************
# *   Copyright (c) 2025 David Carter <dcarter@davidcarter.ca>              *
# *                                                                         *
# *   This program is free software; you can redistribute it and/or modify  *
# *   it under the terms of the GNU Lesser General Public License (LGPL)    *
# *   as published by the Free Software Foundation; either version 2 of     *
# *   the License, or (at your option) any later version.                   *
# *   for detail see the LICENCE text file.                                 *
# *                                                                         *
# *   This program is distributed in the hope that it will be useful,       *
# *   but WITHOUT ANY WARRANTY; without even the implied warranty of        *
# *   MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the         *
# *   GNU Library General Public License for more details.                  *
# *                                                                         *
# *   You should have received a copy of the GNU Library General Public     *
# *   License along with this program; if not, write to the Free Software   *
# *   Foundation, Inc., 59 Temple Place, Suite 330, Boston, MA  02111-1307  *
# *   USA                                                                   *
# *                                                                         *
# ***************************************************************************
"""Class for creating material files from a spreadsheet"""

__title__ = "FreeCAD Materials Generation"
__author__ = "David Carter"
__url__ = "https://www.davesrocketshop.com"

from openpyxl import load_workbook
import os
import uuid
from typing import Any
import cv2
from base64 import b64encode
from PIL import Image
from io import BytesIO
import math

FILENAME = "Resources/Data/Wood Properties FC.xlsx"
IMAGES = "Resources/Data/Images"
OUTPUT_DIR = "Resources/Materials"

ROW_MIN = 5
ROW_MAX = 249
COLUMN_MAX = 45

# Column numbers
COLUMN_NAME = 0 # A
COLUMN_REF1 = 1 # B
COLUMN_IMAGE = 2 # C
COLUMN_ALT_NAMES = 3 # D
COLUMN_SPECIES = 4 # E
COLUMN_REF2 = 5 # F
COLUMN_SOFTWOOD = 6 # G
COLUMN_RANGE = 7 # H
COLUMN_CITES = 8 # I
COLUMN_IUCN_REDLIST = 9 # J
COLUMN_IUCN_REDLIST_URL = 10 # K
COLUMN_STEAM_BEND = 11 # L
COLUMN_HARDNESS = 12 # M
COLUMN_DENSITY = 13 # N
COLUMN_FLEX_MODULUS = 14 # O
COLUMN_SOUND_COEFFICIENT = 15 # P
COLUMN_FLEX_MOD_TANG_LONG = 16 # Q
COLUMN_FLEX_MOD_RAD_LONG = 17 # R
COLUMN_SHEAR_LONG_RAD = 18 # S
COLUMN_SHEAR_LONG_TANG = 19 # T
COLUMN_SHEAR_RAD_TANG = 20 # U
COLUMN_FLEX_STRENGTH = 21 # V
COLUMN_COMPRESS = 22 # W
COLUMN_COMPRESS_STRENGTH_CROSS = 23 # X
COLUMN_SHEAR_LONG = 24 # Y
COLUMN_ULTIMATE_STRENGTH_LONG = 25 # Z
COLUMN_ULTIMATE_STRENGTH_CROSS = 26 # AA
COLUMN_POISSON_LONG_RAD = 27 # AB
COLUMN_POISSON_LONG_TANG = 28 # AC
COLUMN_POISSON_RAD_TANG = 29 # AD
COLUMN_POISSON_TANG_RAD = 30 # AE
COLUMN_POISSON_RAD_LONG = 31 # AF
COLUMN_POISSON_TANG_LONG = 32 # AG
COLUMN_MAX_LOAD = 33 # AH
COLUMN_THERMAL_CONDUCTIVITY = 34 # AI
COLUMN_SHRINK_RAD = 35 # AJ
COLUMN_SHRINK_TAN = 36 # AK
COLUMN_SHRINK_VOL = 37 # AL
COLUMN_LONG_SHRINK = 38 # AM
COLUMN_UUID = 39 # AN
COLUMN_UUID2 = 40 # AO

# Averaged values
VrlBack = 0.056
VlrBack = 0.376
VrlTop = 0.048
VlrTop = 0.386

def parseURL(cell) -> str:
    if cell.hyperlink:
        return str(cell.hyperlink.target)
    return cell.value

def parseSteam(cell) -> str | None:
    if cell.value is None:
        return None
    steam = cell.value
    if steam == '?':
        return None
    return steam
    # return steam.strip("%")

def parseBool(cell) -> bool:
    value = cell.value
    if isinstance(value, bool):
        return value
    if value.lower() in ["true", "=true()",  "1"]:
        return True
    return False

def parseCell(cell : Any) -> tuple[Any, bool]:
    value = cell.value
    if value is None:
        return '', False
    if isinstance(value, str) and value.startswith("="):
        if value == "=VrlBack":
            return VrlBack, True
        if value == "=VlrBack":
            return VlrBack, True
        if value == "=VrlTop":
            return VrlTop, True
        if value == "=VlrTop":
            return VlrTop, True
    return value, False

# def parseFloatCell(cell : str) -> float:
#     try:
#         return float(parseCell(cell))
#     except ValueError:
#         pass
#     return 0

def parseRow(row : tuple) -> dict:
    result = {}
    result["name"] = str(row[COLUMN_NAME].value).strip().title()
    result["softwood"] = parseBool(row[COLUMN_SOFTWOOD])
    result["steam"] = parseSteam(row[COLUMN_STEAM_BEND])
    result["hardness"] = row[COLUMN_HARDNESS].value
    result["density"] = row[COLUMN_DENSITY].value
    result["flex_mod"] = row[COLUMN_FLEX_MODULUS].value
    result["flex_strength"] = row[COLUMN_FLEX_STRENGTH].value
    result["compress"] = row[COLUMN_COMPRESS].value
    result["shrink_rad"] = row[COLUMN_SHRINK_RAD].value
    result["shrink_tan"] = row[COLUMN_SHRINK_TAN].value
    result["shrink_vol"] = row[COLUMN_SHRINK_VOL].value
    result["image"] = row[COLUMN_IMAGE].value
    result["species"] = row[COLUMN_SPECIES].value
    result["alt"] = row[COLUMN_ALT_NAMES].value
    result["ref1"] = parseURL(row[COLUMN_REF1])
    result["ref2"] = parseURL(row[COLUMN_REF2])
    if row[COLUMN_UUID].value is None:
        row[COLUMN_UUID].value = str(uuid.uuid4())
    result["UUID"] = row[COLUMN_UUID].value
    result["UUID2"] = row[COLUMN_UUID2].value
    result["range"] = row[COLUMN_RANGE].value
    result["CITES"] = row[COLUMN_CITES].value
    result["Redlist"] = row[COLUMN_IUCN_REDLIST].value
    result["RedlistURL"] = row[COLUMN_IUCN_REDLIST_URL].value
    result["FlexModulusTangLong"] = row[COLUMN_FLEX_MOD_TANG_LONG].value
    result["FlexModulusRadLong"] = row[COLUMN_FLEX_MOD_RAD_LONG].value
    result["ShearLongRad"] = row[COLUMN_SHEAR_LONG_RAD].value
    result["ShearLongTang"] = row[COLUMN_SHEAR_LONG_TANG].value
    result["ShearRadTang"] = row[COLUMN_SHEAR_RAD_TANG].value
    result["UltimateLong"] = row[COLUMN_ULTIMATE_STRENGTH_LONG].value
    result["UltimateCross"] = row[COLUMN_ULTIMATE_STRENGTH_CROSS].value
    result["CompressCross"] = row[COLUMN_COMPRESS_STRENGTH_CROSS].value
    result["ShearLong"] = row[COLUMN_SHEAR_LONG].value
    result["PoissonLongRad"] = row[COLUMN_POISSON_LONG_RAD].value
    result["PoissonLongTang"] = row[COLUMN_POISSON_LONG_TANG].value
    result["PoissonRadTang"] = row[COLUMN_POISSON_RAD_TANG].value
    result["PoissonTangRad"] = row[COLUMN_POISSON_TANG_RAD].value
    result["PoissonRadLong"] = row[COLUMN_POISSON_RAD_LONG].value
    result["PoissonTangLong"] = row[COLUMN_POISSON_TANG_LONG].value
    result["ThermalConductivity"] = row[COLUMN_THERMAL_CONDUCTIVITY].value
    result["SoundCoefficient"] = row[COLUMN_SOUND_COEFFICIENT].value
    result["MaxLoad"] = row[COLUMN_MAX_LOAD].value
    return result

def getTags(row : dict) -> list:
    tags = []
    if row['alt']:
        names = row['alt'].split(',')
        for name in names:
            tag = name.strip().lower()
            tags.append(tag)
    # if row['tags'] is not None:
    #     names = row['tags'].split(',')
    #     for name in names:
    #         tag = name.strip().lower()
    #         tags.append(tag)
    return tags

def getRange(row : dict) -> list:
    ranges = []
    names = row['range'].split(',')
    for range in names:
        ranges.append(range.strip().upper())
    return ranges

def createInherits(row : dict) -> str:
    yam = ""
    if row["softwood"]:
        yam =  'Inherits:\n'
        yam += '  Softwood:\n'
        yam += '    UUID: "f9d83964-24ca-44df-a570-d1af36756a99"\n'
    else:
        yam =  'Inherits:\n'
        yam += '  Hardwood:\n'
        yam += '    UUID: "2a78c735-c21b-4bf9-8606-149d74f88fa8"\n'
    return yam

def createBotanical(row : dict) -> str:
    yam =   "  Wood - Botanical:\n"
    yam +=  '    UUID: "1273eaa6-8185-4130-8072-ff61132568d9"\n'
    if row["species"]:
        yam += f'    Species: "{row["species"].strip().title()}"\n'
    if row["ref2"]:
        yam += f'    SpeciesURL: "{row["ref2"]}"\n'
    if row["ref1"]:
        yam += f'    WoodDatabase: "{row["ref1"]}"\n'
    if row["softwood"] is not None:
        yam += f'    Softwood: "{row["softwood"]}"\n'
    ranges = getRange(row)
    if ranges:
        yam += f'    Range:\n'
        for range in ranges:
            yam += f'      - "{range}"\n'
    if row["CITES"]:
        yam += f'    CITESAppendix: "{row["CITES"]}"\n'
    if row["Redlist"]:
        yam += f'    IUCNRedList: "{row["Redlist"]}"\n'
    if row["RedlistURL"]:
        yam += f'    IUCNRedListURL: "{row["RedlistURL"]}"\n'
    return yam

def createMachinability(row : dict) -> str:
    if row["MachSpeedHSS"] or row["MachSpeedCarbide"] or row["MachUnitCuttingForce"] or row["MachChipThickness"]:
        yam =   '  Machinability:\n'
        yam +=  '    UUID: "9d81fcb2-bf81-48e3-bb57-d45ecf380096"\n'
        if row["MachSpeedHSS"]:
            yam += f'    SurfaceSpeedHSS: "{row["MachSpeedHSS"]} mm/min"\n'
        if  row["MachSpeedCarbide"]:
            yam += f'    SurfaceSpeedCarbide: "{row["MachSpeedCarbide"]} mm/min"\n'
        if row["MachUnitCuttingForce"]:
            yam += f'    UnitCuttingForce: "{row["MachUnitCuttingForce"]} N/mm^2"\n'
        if row["MachChipThickness"]:
            yam += f'    ChipThicknessExponent: "{row["MachChipThickness"]}"\n'
        return yam
    else:
        return ""

def createHardness(row : dict) -> str:
    if row["hardness"]:
        yam =   '  Hardness:\n'
        yam +=  '    UUID: "3d1a6141-d032-4d82-8bb5-a8f339fff8ad"\n'
        yam += f'    Hardness: "{row["hardness"]}"\n'
        yam += f'    HardnessUnits: "N"\n'
        return yam
    else:
        return ""

def createShrinkage(row : dict) -> str:
    yam =   '  Wood - Shrinkage:\n'
    yam +=  '    UUID: "ec84f5bb-99cf-448a-86a5-cac2ebcab31c"\n'
    shrinkRadial = row["shrink_rad"]
    shrinkTangential = row["shrink_tan"]
    shrinkVolume = row["shrink_vol"]
    if shrinkRadial:
        yam += f'    ShrinkRadial: "{shrinkRadial * 100.0}"\n'
    if shrinkTangential:
        yam += f'    ShrinkTangential: "{shrinkTangential * 100.0}"\n'
    if shrinkVolume:
        yam += f'    ShrinkVolume: "{shrinkVolume * 100.0}"\n'
    if shrinkRadial and shrinkTangential and shrinkVolume:
        shrinkLong = max(1 - (1 - shrinkVolume)/((1 - shrinkRadial)*(1 - shrinkTangential)), 0)
        yam += f'    ShrinkLong: "{shrinkLong * 100.0}"\n'
    return yam

def createThermal(row : dict) -> str:
    if row["ThermalConductivity"]:
        yam =   '  Thermal:\n'
        yam +=  '    UUID: "9959d007-a970-4ea7-bae4-3eb1b8b883c7"\n'
        yam += f'    ThermalConductivity: "{row["ThermalConductivity"]} W/m/K"\n'
        return yam
    else:
        return ""

def createSound(row : dict) -> str:
    density = row["density"]
    young = row["flex_mod"]
    if density and young:
        coefficient = math.sqrt(young * 1e6 / math.pow(density, 3))
        yam =   '  Sound:\n'
        yam +=  '    UUID: "6b7f44ab-e48d-4568-98aa-1d88a8b6e57d"\n'
        yam += f'    SoundRadiationCoefficient: "{coefficient:.1f} m^4/kg/s"\n'
        return yam
    return ""

def createLinearElastic(row : dict) -> str:
    yam =   '  LinearElastic:\n'
    yam +=  '    UUID: "7b561d1d-fb9b-44f6-9da9-56a4f74d7536"\n'
    yam += f'    Density: "{row["density"]} kg/m^3"\n'

    # This produces inconsistent results
    # if row["flex_mod"] and row["ShearLong"]:
    #     yam += f'    PoissonRatio: "{(row["flex_mod"] * 1000.0 / (2.0 * row["ShearLong"])) - 1.0:.3f}"\n'

    # if row["ShearLong"]:
    #     yam += f'    ShearModulus: "{row["ShearLong"]} kPa"\n'
    if row["flex_mod"]:
        yam += f'    YoungsModulus: "{row["flex_mod"]} MPa"\n'
    if row["compress"]:
        yam += f'    CompressiveStrength: "{row["compress"]} kPa"\n'
    if row["UltimateLong"]:
        yam += f'    UltimateTensileStrength: "{row["UltimateLong"]} kPa"\n'

    return yam

def createWood(row : dict) -> str:
    yam =   '  Wood:\n'
    yam +=  '    UUID: "901459aa-fd5e-43b8-aad6-71578f76c3f6"\n'
    # yam += f'    MoistureContent: "{row["density"]}"\n'
    if row["steam"]:
        yam += f'    SteamBendable: "{row["steam"] * 100.0}"\n'
    yam += f'    Density: "{row["density"]} kg/m^3"\n'
    if row["PoissonLongRad"]:
        yam += f'    PoissonRatioLongRad: "{row["PoissonLongRad"]}"\n'
    if row["PoissonLongTang"]:
        yam += f'    PoissonRatioLongTan: "{row["PoissonLongTang"]}"\n'
    if row["PoissonRadTang"]:
        yam += f'    PoissonRatioRadTan: "{row["PoissonRadTang"]}"\n'
    if row["PoissonTangRad"]:
        yam += f'    PoissonRatioTanRad: "{row["PoissonTangRad"]}"\n'
    if row["PoissonRadLong"]:
        yam += f'    PoissonRatioRadLong: "{row["PoissonRadLong"]}"\n'
    if row["PoissonTangLong"]:
        yam += f'    PoissonRatioTanLong: "{row["PoissonTangLong"]}"\n'
    if row["ShearLong"]:
        yam += f'    ShearStrengthLong: "{row["ShearLong"]} kPa"\n'
    if row["ShearLongRad"] and row["flex_mod"]:
        yam += f'    ShearModulusLongRad: "{row["ShearLongRad"] * row["flex_mod"]:.2f} MPa"\n'
    if row["ShearLongTang"] and row["flex_mod"]:
        yam += f'    ShearModulusLongTan: "{row["ShearLongTang"] * row["flex_mod"]:.2f} MPa"\n'
    if row["ShearRadTang"] and row["flex_mod"]:
        yam += f'    ShearModulusRadTan: "{row["ShearRadTang"] * row["flex_mod"]:.2f} MPa"\n'
    if row["flex_mod"]:
        yam += f'    YoungsModulusLong: "{row["flex_mod"]} MPa"\n'
    if row["FlexModulusTangLong"] and row["flex_mod"]:
        yam += f'    YoungsModulusTanLong: "{row["FlexModulusTangLong"] * row["flex_mod"]:.2f} MPa"\n'
    if row["FlexModulusRadLong"] and row["flex_mod"]:
        yam += f'    YoungsModulusRadLong: "{row["FlexModulusRadLong"] * row["flex_mod"]:.2f} MPa"\n'
    if row["UltimateLong"]:
        yam += f'    UltimateStrengthLong: "{row["UltimateLong"]} kPa"\n'
    if row["UltimateCross"]:
        yam += f'    UltimateStrengthCross: "{row["UltimateCross"]} kPa"\n'
    if row["compress"]:
        yam += f'    CompressiveStrengthLong: "{row["compress"]} kPa"\n'
    if row["CompressCross"]:
        yam += f'    CompressiveStrengthCross: "{row["CompressCross"]} kPa"\n'
    if row["flex_strength"]:
        yam += f'    ModulusOfRuptureLong: "{row["flex_strength"]} kPa"\n'
    if row["MaxLoad"]:
        yam += f'    WorkToMaximumLoad: "{row["MaxLoad"]} kJ/m^3"\n'
    return yam

def createAppearance(base : str | None, diffuse : tuple) ->str:
    yam =   "AppearanceModels:\n"
    yam +=  "  Texture Rendering:\n"
    yam +=  '    UUID: "bbdcc65b-67ca-489c-bd5c-a36e33d1c160"\n'
    yam +=  '    AmbientColor: "(0.333333, 0.333333, 0.333333, 1)"\n'
    yam += f'    DiffuseColor: "{diffuse}"\n'
    yam +=  '    EmissiveColor: "(0, 0, 0, 1)"\n'
    yam +=  '    Shininess: "0.9"\n'
    yam +=  '    SpecularColor: "(0.533333, 0.533333, 0.533333, 1)"\n'
    yam +=  '    Transparency: "0"\n'
    if base is not None:
        yam += '    TextureImage:' + base
    return yam

def createYaml(row : dict, base : str | None, diffuse : tuple, averaged : bool = False) -> str:
    yam = "# File created by the Woods workbench\n"
    yam += "General:\n"
    # Add UUIDs
    if averaged:
        yam += f'  UUID: "{row["UUID2"]}"\n'
        yam += f'  Name: "{row["name"]} (Averaged)"\n'
    else:
        yam += f'  UUID: "{row["UUID"]}"\n'
        yam += f'  Name: "{row["name"]}"\n'
    yam += f'  Author: "Gregory Holmberg"\n'
    yam += f'  License: "CDLA-Sharing-1.0"\n'
    yam += f'  SourceURL: "https://research.fs.usda.gov/treesearch/62200"\n'
    yam += f'  ReferenceSource: "USDA FPL Wood Handbook 2021"\n'
    tags = getTags(row)
    if len(tags) > 0:
        yam += f'  Tags:\n'
        for tag in tags:
            yam += f'    - "{tag}"\n'
    yam += f'  Description: >-2\n'
    yam += '    Automatically created by the Woods workbench.\n'
    if averaged:
        yam += '    \n'
        yam += '    \n'
        yam += '    This file includes averaged values in the absence of known values.\n'
        yam += '    Use with caution as the values may produce incorrect results.\n'
    yam += createInherits(row)

    yam += 'Models:\n'
    yam += createBotanical(row)
    # yam += createMachinability(row)
    yam += createHardness(row)
    yam += createShrinkage(row)
    yam += createThermal(row)
    yam += createSound(row)
    yam += createLinearElastic(row) #- produces bad results for poisson ratio
    yam += createWood(row)
    yam += createAppearance(base, diffuse)

    return yam

def createCard(row : dict, base : str | None, diffuse : tuple) -> None:
    name = row["name"]
    if name is not None:
        # if row["averaged"]:
        #     yaml = createYaml(row, base, diffuse, True)
        #     outputName = f"{OUTPUT_DIR}/{name} (Averaged).FCMat"
        #     outfile = open(outputName, "w", encoding="utf-8")
        #     outfile.write(yaml)
        #     outfile.close()

        yaml = createYaml(row, base, diffuse, False)
        outputName = f"{OUTPUT_DIR}/{name}.FCMat"
        outfile = open(outputName, "w", encoding="utf-8")
        outfile.write(yaml)
        outfile.close()

def imageToPng(imageData : bytes) -> bytes:
    # Create an in-memory binary stream for the input JPG data
    imageBuffer = BytesIO(imageData)

    # Open the image using Pillow
    img = Image.open(imageBuffer)

    # Create an in-memory binary stream for the output PNG data
    pngBuffer = BytesIO()

    # Save the image as PNG to the output buffer
    img.save(pngBuffer, format="PNG")

    # Get the bytes data of the PNG image
    pngData = pngBuffer.getvalue()

    return pngData

def checkImage(data : dict) -> tuple[str | None, Any]:
    base = None
    diffuse = (0.859, 0.780, 0.584, 1)
    if data["image"] is not None:
        image = f"{IMAGES}/{data['image']}"
        if os.path.exists(image):
            im = cv2.imread(image)
            A = cv2.mean(im)

            # BGR to RGB
            diffuse = (A[2] / 255.0, A[1] / 255.0, A[0] / 255.0, 1.0)

            # Convert the image to base64
            with open(image, "rb") as image_file:
                # v1.0 only works with PNG. Use this to maintain compatibility
                png = imageToPng(image_file.read())
                encoded_string = b64encode(png)
                encoded_output = encoded_string.decode('utf-8')

            base = " |-2"
            while len(encoded_output) > 0:
                base += "\n      "
                base += encoded_output[:74]
                encoded_output = encoded_output[74:]
            base += "\n"
        else:
            print(f"Missing image '{image}'")

    return base, diffuse

# Create the output folder if required
os.makedirs(OUTPUT_DIR, exist_ok=True)

# wb = load_workbook(filename=FILENAME, read_only=True)
wb = load_workbook(filename=FILENAME, read_only=False)
ws = wb['All']
# for row in ws.iter_rows(min_row=5, max_row=247, max_col=22, values_only=True):
for row in ws.iter_rows(min_row=ROW_MIN, max_row=ROW_MAX, max_col=COLUMN_MAX):
    cell = row[20]
    parsed = parseRow(row)
    base, diffuse = checkImage(parsed)
    createCard(parsed, base, diffuse)

wb.save(filename=FILENAME)
